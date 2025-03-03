from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import pytz 

# Путь к Excel файлу
EXCEL_FILE_PATH = "orders/orders.xlsx"

timezone = pytz.timezone("Europe/Moscow")

# Функция для добавления данных о заказе в Excel
async def add_order_to_excel(order_data):
    # Проверяем, существует ли файл
    if not os.path.exists(EXCEL_FILE_PATH):
        # Если файл не существует, создаём новый и записываем заголовки
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        headers = ["Order ID", "User ID", "Total Price", "Delivery Data", "Date"]
        ws.append(headers)
        wb.save(EXCEL_FILE_PATH)
    else:
        # Если файл существует, загружаем его
        wb = load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
    
    # Преобразуем дату, если это объект datetime
    order_date = order_data["date"]
    if isinstance(order_date, datetime):
        if order_date.tzinfo:
            # Переводим в нужную временную зону
            order_date = order_date.astimezone(timezone)  # Переводим в МСК (+3)
        else:
            # Если нет информации о временной зоне, просто устанавливаем её
            order_date = timezone.localize(order_date)  # Устанавливаем временную зону вручную

        formatted_date = order_date.strftime("%Y-%m-%d %H:%M:%S")
    else:
        formatted_date = str(order_date)  # Если это не datetime, просто преобразуем в строку

    # Добавляем данные о заказе
    order_row = [
        order_data["order_id"],
        order_data["user_id"],
        order_data["total_price"],
        order_data["delivery_data"],
        formatted_date  # Форматируем дату в строку
    ]
    ws.append(order_row)

    # Сохраняем файл
    wb.save(EXCEL_FILE_PATH)
